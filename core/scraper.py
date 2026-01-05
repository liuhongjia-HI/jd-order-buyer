import json
import time
import os
import sys
import random
import re
from datetime import datetime
import threading
from pathlib import Path
from urllib.parse import urljoin
import io
import requests
from playwright.sync_api import sync_playwright, TimeoutError
from playwright_stealth import Stealth
from loguru import logger
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def _data_base_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent.parent

class JDScraper:
    def __init__(self, headless=False):
        self.headless = headless
        self.browser = None
        self.context = None
        self.page = None
        self.detail_page = None
        self.playwright = None
        self.base_dir = _data_base_dir()
        self.auth_file = str(self.base_dir / "auth.json")
        self.profile_name = (os.getenv("JD_PROFILE", "default") or "default").strip()
        default_profile_dir = self.base_dir / "profiles" / self.profile_name
        self.profile_dir = Path(os.getenv("JD_PROFILE_DIR", default_profile_dir)).expanduser().resolve()
        self.fingerprint_file = Path(
            os.getenv("JD_FINGERPRINT_FILE", self.profile_dir / "fingerprint.json")
        ).expanduser().resolve()
        self.use_persistent_context = os.getenv("JD_PERSISTENT_PROFILE", "1") != "0"
        # 可配置下载目录与嵌入图片开关
        self.download_dir = Path(os.getenv("JD_DOWNLOAD_DIR", self.base_dir / "downloads")).expanduser().resolve()
        self.base_url = "https://order.jd.com/center/list.action"
        self.stealth = Stealth()
        self._lock = threading.RLock()
        self.address_cache = {}
        self.embed_images = os.getenv("JD_EMBED_IMAGES", "1") != "0"
        self.goto_retries = self._safe_int(os.getenv("JD_GOTO_RETRIES", "3"), default=3)
        self.risk_wait_s = self._safe_int(os.getenv("JD_RISK_WAIT", "120"), default=120)
        self.rate_limits = {
            "page": self._safe_float(os.getenv("JD_RATE_PAGE_MIN", "1.8"), default=1.8),
            "detail": self._safe_float(os.getenv("JD_RATE_DETAIL_MIN", "2.2"), default=2.2),
            "image": self._safe_float(os.getenv("JD_RATE_IMAGE_MIN", "0.6"), default=0.6),
        }
        self.rate_last = {"page": 0.0, "detail": 0.0, "image": 0.0}
        self.rate_multipliers = {"page": 1.0, "detail": 1.0, "image": 1.0}
        self.rate_backoff_max = self._safe_float(os.getenv("JD_RATE_BACKOFF_MAX", "6"), default=6.0)
        self.image_retries = self._safe_int(os.getenv("JD_IMAGE_RETRIES", "2"), default=2)
        self.browse_prob = self._safe_float(os.getenv("JD_BROWSE_PROB", "0.35"), default=0.35)
        self.browse_every = self._safe_int(os.getenv("JD_BROWSE_EVERY", "0"), default=0)
        self.detail_browse_prob = self._safe_float(os.getenv("JD_DETAIL_BROWSE_PROB", "0.6"), default=0.6)
        self.browse_urls = self._parse_browse_urls(os.getenv("JD_BROWSE_URLS", "https://www.jd.com/,https://home.jd.com/"))
        self.http = requests.Session()
        self.risk_url_keywords = (
            "passport.jd.com",
            "safe.jd.com",
            "risk",
            "captcha",
            "verify",
            "challenge",
        )
        self.risk_text_keywords = (
            "验证码",
            "安全验证",
            "访问过于频繁",
            "异常访问",
            "风险",
            "滑块",
            "请使用京东客户端",
        )
        self.fingerprint = self._load_or_create_fingerprint()
        # Pool of UAs；可通过 JD_UA 固定，优先保持一致性
        self.user_agents = [self.fingerprint["user_agent"]]
        self.locale = self.fingerprint["locale"]
        self.timezone_id = self.fingerprint["timezone_id"]
        self.accept_language = self.fingerprint["accept_language"]
        self.viewport = self.fingerprint["viewport"]
        self.device_scale_factor = self.fingerprint["device_scale_factor"]
        self.is_mobile = self.fingerprint["is_mobile"]

    def _random_sleep(self, min_s=1.5, max_s=4.0):
        """Random delay to mimic human behavior"""
        delay = random.uniform(min_s, max_s)
        time.sleep(delay)

    def _rate_limit(self, kind: str):
        min_interval = self.rate_limits.get(kind, 0)
        if min_interval <= 0:
            return
        multiplier = self.rate_multipliers.get(kind, 1.0)
        min_interval = min_interval * multiplier
        now = time.monotonic()
        last = self.rate_last.get(kind, 0.0)
        sleep_s = min_interval - (now - last)
        if sleep_s > 0:
            time.sleep(sleep_s + random.uniform(0.05, 0.25))
        self.rate_last[kind] = time.monotonic()

    def _bump_backoff(self, kind: str, factor: float = 1.6):
        current = self.rate_multipliers.get(kind, 1.0)
        next_val = min(current * factor, self.rate_backoff_max)
        self.rate_multipliers[kind] = next_val

    def _decay_backoff(self, kind: str, decay: float = 0.92):
        current = self.rate_multipliers.get(kind, 1.0)
        if current > 1.0:
            self.rate_multipliers[kind] = max(1.0, current * decay)

    def _humanize_page(self):
        """轻量行为模拟：使用脚本滚动，避免占用真实鼠标。"""
        try:
            self.page.wait_for_timeout(random.randint(300, 900))
            delta = random.randint(300, 900)
            self.page.evaluate(
                """(d) => { window.scrollBy({top: d, behavior: 'smooth'}); }""",
                delta
            )
        except Exception:
            pass

    def _dwell_and_scroll(self, page, min_s=1.0, max_s=2.6):
        try:
            time.sleep(random.uniform(min_s, max_s))
            for _ in range(random.randint(1, 2)):
                delta = random.randint(300, 900)
                page.evaluate(
                    """(d) => { window.scrollBy({top: d, behavior: 'smooth'}); }""",
                    delta
                )
                page.wait_for_timeout(random.randint(300, 900))
            if random.random() < 0.4:
                page.evaluate("""() => { window.scrollTo({top: 0, behavior: 'smooth'}); }""")
        except Exception:
            pass

    def _simulate_browse_path(self, stage="start"):
        if not self.context or not self.browse_urls:
            return
        if self.browse_prob <= 0:
            return
        if random.random() > self.browse_prob:
            return

        browse_page = self.context.new_page()
        try:
            try:
                self.stealth.apply_stealth_sync(browse_page)
            except Exception:
                pass
            logger.info(f"Simulating browse path ({stage})...")
            for url in self.browse_urls:
                try:
                    self._rate_limit("page")
                    browse_page.goto(url, wait_until="domcontentloaded", timeout=20000)
                    try:
                        browse_page.wait_for_load_state("networkidle", timeout=8000)
                    except Exception:
                        pass
                    reason = self._detect_risk_page(browse_page)
                    if reason and self._handle_risk_page(browse_page, reason, fatal=False):
                        break
                    self._dwell_and_scroll(browse_page, min_s=1.0, max_s=2.6)
                    self._decay_backoff("page")
                except Exception as e:
                    logger.warning(f"Browse path step failed: {url} -> {e}")
                    self._bump_backoff("page")
        finally:
            try:
                browse_page.close()
            except Exception:
                pass

    def _detect_risk_page(self, page):
        url = (page.url or "").lower()
        for kw in self.risk_url_keywords:
            if kw in url:
                return f"url:{kw}"

        try:
            title = page.title() or ""
        except Exception:
            title = ""

        try:
            text = page.evaluate(
                "() => (document.body && document.body.innerText) ? document.body.innerText.slice(0, 2000) : ''"
            ) or ""
        except Exception:
            text = ""

        haystack = f"{title}\n{text}"
        for kw in self.risk_text_keywords:
            if kw in haystack:
                return f"text:{kw}"
        return ""

    def _handle_risk_page(self, page, reason: str, fatal: bool = True):
        if not reason:
            return False
        logger.warning(f"Detected risk page ({reason}).")
        self._bump_backoff("page", factor=2.0)
        self._bump_backoff("detail", factor=2.0)
        wait_s = max(5, self.risk_wait_s)
        if self.headless:
            if fatal:
                raise Exception(f"检测到风控/验证码页，请人工处理后重试: {reason}")
            return True

        start = time.time()
        while time.time() - start < wait_s:
            time.sleep(2)
            if not self._detect_risk_page(page):
                logger.success("Risk page cleared manually.")
                return False

        if fatal:
            raise Exception(f"风控页面未解除，请稍后重试: {reason}")
        return True

    def _get_detail_page(self):
        if self.detail_page and not self.detail_page.is_closed():
            return self.detail_page
        self.detail_page = self.context.new_page()
        try:
            self.stealth.apply_stealth_sync(self.detail_page)
        except Exception:
            pass
        return self.detail_page

    def _reset_detail_page(self):
        if self.detail_page and not self.detail_page.is_closed():
            try:
                self.detail_page.close()
            except Exception:
                pass
        self.detail_page = None

    def _open_jd_home(self, retries=2):
        """Open JD homepage with basic retry to avoid staying on about:blank."""
        for attempt in range(retries):
            try:
                self._rate_limit("page")
                url = f"https://www.jd.com/?r={int(time.time()*1000)}"
                logger.info(f"Opening JD homepage (attempt {attempt+1})...")
                # Use domcontentloaded which is faster and sufficient for warm-up
                self.page.goto(url, wait_until="domcontentloaded", timeout=45000)
                
                # Try waiting for network idle but ignore timeout if page is usable
                try:
                    self.page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass

                # small human-like pause
                self._random_sleep(1.2, 2.5)
                reason = self._detect_risk_page(self.page)
                if reason:
                    if self._handle_risk_page(self.page, reason, fatal=False):
                        continue
                if "jd.com" in self.page.url:
                    return True
            except Exception as e:
                logger.warning(f"JD homepage open failed (attempt {attempt+1}): {e}")
        return False

    def _ensure_auth_state(self):
        """Ensure we have a stored login state before scraping."""
        if not os.path.exists(self.auth_file):
            raise Exception("未检测到 auth.json，请先登录一次再尝试采集。")

    def _wait_for_auth_cookie(self, timeout=300000):
        """Wait until JD login cookies appear (pt_key/pt_pin or user profile)."""
        deadline = time.time() + (timeout / 1000)
        while time.time() < deadline:
            try:
                cookies = self.context.cookies()
                names = {c.get("name") for c in cookies}
                # Relaxed cookie check: pt_key/pin/thor or even just 'nickname' which often appears after login
                if any(n in names for n in ("pt_key", "pin", "thor", "pwdt_id", "unick")):
                    return True
            except Exception:
                pass
            
            # Check URL for successful login redirects
            current_url = self.page.url
            if any(s in current_url for s in ("order.jd.com", "home.jd.com", "user.jd.com", "joycenter.jd.com")):
                 return True
            
            # If we are strictly on www.jd.com, check if we have username element (soft check)
            if "www.jd.com" in current_url:
                try:
                    # Often when logged in on homepage, nickname appears
                    if self.page.query_selector(".nickname") or self.page.query_selector(".user_name"):
                         return True
                except:
                    pass

            time.sleep(1.0)
        raise TimeoutError("等待登录 Cookie 超时")


    def _open_order_after_login(self, retries=2):
        """Try to land on订单列表 after login cookies are detected."""
        for i in range(retries):
            try:
                self._goto_with_retry(f"{self.base_url}?s=4096", wait_until="domcontentloaded", retries=2)
            except Exception as nav_err:
                logger.warning(f"订单页跳转失败重试: {nav_err}")
            if "passport.jd.com" not in self.page.url:
                return True
            logger.warning("登录后仍在登录页，重试跳转订单列表...")
            time.sleep(2)
        return False

    def start_browser(self, use_storage: bool = True):
        logger.info(f"Launching Browser (Headless={self.headless})...")
        self.playwright = sync_playwright().start()
        # Removed global hook to prevent potential startup hangs
        if not self.http:
            self.http = requests.Session()

        ua = self.user_agents[0]
        # Try launch options: Bundled -> Edge -> Chrome
        # Try launch options: Bundled -> Edge -> Chrome
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--no-first-run",
            "--no-service-autorun",
            "--password-store=basic",
            "--use-mock-keychain",
        ]

        # Load state if exists
        load_options = {"storage_state": self.auth_file} if (use_storage and os.path.exists(self.auth_file)) else {}
        context_options = {
            "user_agent": ua,
            "viewport": {"width": self.viewport["width"], "height": self.viewport["height"]},
            "locale": self.locale,
            "timezone_id": self.timezone_id,
            "device_scale_factor": self.device_scale_factor,
            "is_mobile": self.is_mobile,
            "extra_http_headers": {
                "Accept-Language": self.accept_language,
                "Referer": "https://www.jd.com/",
                "Origin": "https://www.jd.com"
            },
        }

        if self.use_persistent_context:
            try:
                self.profile_dir.mkdir(parents=True, exist_ok=True)
                logger.info(f"Launching persistent context: {self.profile_dir}")
                try:
                    self.context = self.playwright.chromium.launch_persistent_context(
                        user_data_dir=str(self.profile_dir),
                        headless=self.headless,
                        args=launch_args,
                        ignore_default_args=["--enable-automation"],
                        **context_options,
                        **load_options,
                    )
                except TypeError:
                    # Older Playwright may not accept storage_state here.
                    self.context = self.playwright.chromium.launch_persistent_context(
                        user_data_dir=str(self.profile_dir),
                        headless=self.headless,
                        args=launch_args,
                        ignore_default_args=["--enable-automation"],
                        **context_options,
                    )
                self.browser = self.context.browser
            except Exception as e:
                logger.warning(f"Persistent context launch failed, fallback to normal context: {e}")
                self.use_persistent_context = False

        if not self.use_persistent_context:
            try:
                logger.info("Generic launch...")
                self.browser = self.playwright.chromium.launch(
                    headless=self.headless,
                    args=launch_args,
                    ignore_default_args=["--enable-automation"]
                )
            except Exception:
                try:
                    logger.info("Bundled browser not found. Trying system Edge...")
                    self.browser = self.playwright.chromium.launch(
                        channel="msedge",
                        headless=self.headless,
                        args=launch_args,
                        ignore_default_args=["--enable-automation"]
                    )
                except Exception:
                    logger.info("Edge not found. Trying system Chrome...")
                    self.browser = self.playwright.chromium.launch(
                        channel="chrome",
                        headless=self.headless,
                        args=launch_args,
                        ignore_default_args=["--enable-automation"]
                    )

            self.context = self.browser.new_context(
                **context_options,
                **load_options,
            )
        
        # Inject stealth scripts to hide webdriver property
        self.context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            Object.defineProperty(navigator, 'languages', {
                get: () => ['zh-CN', 'zh', 'en']
            });
            const originalQuery = window.navigator.permissions.query;
            window.navigator.permissions.query = (parameters) => (
                parameters.name === 'notifications' ?
                Promise.resolve({ state: Notification.permission }) :
                originalQuery(parameters)
            );
        """)

        self.page = self.context.new_page()
        try:
            self.stealth.apply_stealth_sync(self.context)
            self.stealth.apply_stealth_sync(self.page)
        except Exception as e:
            logger.warning(f"Stealth apply failed: {e}")
        # Tighter but consistent timeouts avoid long hangs while staying human-like
        self.context.set_default_timeout(15000)
        self.context.set_default_navigation_timeout(20000)

    def close_browser(self):
        if self.detail_page:
            try:
                self.detail_page.close()
            except Exception:
                pass
            self.detail_page = None
        if self.context:
            self.context.close()
            self.context = None
        if self.browser:
            try:
                self.browser.close()
            except Exception:
                pass
            self.browser = None
        if self.playwright:
            self.playwright.stop()
            self.playwright = None
        self.page = None
        if self.http:
            try:
                self.http.close()
            except Exception:
                pass
            self.http = None

    def login(self, force_fresh: bool = False):
        """Manually login and save state."""
        with self._lock:
            return self._login_locked(force_fresh=force_fresh)

    def _login_locked(self, force_fresh: bool = False):
        logger.info("Starting login process (Stealth Mode)...")
        if not self.browser:
            self.headless = False
            self.start_browser(use_storage=not force_fresh)
        
        try:
            # If已有会话，直接打开订单列表
            if os.path.exists(self.auth_file) and not force_fresh:
                logger.info("检测到已有会话，直接打开订单列表...")
                self.page.goto(f"{self.base_url}?s=4096", wait_until="domcontentloaded")
                self.page.wait_for_load_state("networkidle")
                if "passport.jd.com" not in self.page.url:
                    logger.success("会话有效，已打开我的订单列表。")
                    self.context.storage_state(path=self.auth_file)
                    return True
                logger.warning("会话已失效，转为扫码登录（不再复用旧存储状态）。")
                # Drop stale cookies/storage to avoid被重定向到风控或验证码页
                self.close_browser()
                self.headless = False
                self.start_browser(use_storage=False)

            # 先打开首页伪装正常访问，再跳转登录页
            logger.info("Opening JD homepage before login for warm-up...")
            opened = self._open_jd_home()
            if not opened:
                logger.warning("JD 首页多次打开失败，仍将尝试登录页。")

            self.page.goto("https://passport.jd.com/new/login.aspx", wait_until="domcontentloaded")
            logger.info("Please scan the QR code to login...")

            # Wait for auth cookies to appear
            self._wait_for_auth_cookie(timeout=300000)
            logger.info("检测到登录 Cookie，跳转订单页确认会话...")

            if not self._open_order_after_login():
                raise Exception("登录未生效，请重试扫码。")

            logger.success("Login detected! Saving state and continuing.")
            self.context.storage_state(path=self.auth_file)
            return True
        except Exception as e:
            logger.error(f"Login failed: {e}")
            return False
        finally:
            self.close_browser()

    def scrape_orders(self, year_filter="1"):
        """
        Robust sequential scraping.
        """
        with self._lock:
            return self._scrape_locked(year_filter)

    def _scrape_locked(self, year_filter="1"):
        logger.info(f"Starting robust scrape task. Filter d={year_filter}")
        
        # Auto-trigger login if no auth is present
        if not os.path.exists(self.auth_file):
            logger.warning("auth.json 未找到，自动弹出浏览器进行扫码登录...")
            login_success = self.login()
            if not login_success:
                return {"status": "error", "message": "登录失败或超时，请扫码完成后再试。"}

        if not self.browser:
            self.start_browser()

        orders = []
        page_num = 1
        max_retries = 3
        reauth_attempted = False
        
        try:
            # Initial Navigation
            url = f"{self.base_url}?d={year_filter}&s=4096"
            self._simulate_browse_path(stage="start")
            self._goto_with_retry(url, wait_until="domcontentloaded")
            self.page.wait_for_load_state("networkidle")

            while True:
                logger.info(f"Processing Page {page_num}...")
                self._humanize_page()
                
                # Check for auth redirect
                if "passport.jd.com" in self.page.url:
                    if not reauth_attempted:
                        logger.warning("Session expired, auto re-login once...")
                        reauth_attempted = True
                        # Close current browser/context and refresh auth
                        self.close_browser()
                        login_ok = self.login(force_fresh=True)
                        if not login_ok:
                            raise Exception("Session expired and re-login failed.")
                        # Fresh browser with new storage state
                        if self.browser:
                            self.close_browser()
                        self.start_browser()
                        self._goto_with_retry(url, wait_until="domcontentloaded")
                        self.page.wait_for_load_state("networkidle")
                        continue
                    raise Exception("Session expired. Please re-login.")
                
                # Retry logic for current page parsing
                retry_count = 0
                success = False
                last_first_id = None
                
                while retry_count < max_retries:
                    try:
                        reason = self._detect_risk_page(self.page)
                        if reason:
                            self._handle_risk_page(self.page, reason, fatal=True)
                        self._wait_for_orders_ready()
                        # Use correct selector for order tbodies
                        rows = self.page.query_selector_all("tbody[id^='tb-']")
                        
                        if not rows:
                            raise Exception("页面没有找到订单列表")

                        logger.info(f"Found {len(rows)} order entries on page {page_num}.")

                        for row in rows:
                            items = self._parse_row(row)
                            if items:
                                orders.extend(items)

                        last_first_id = rows[0].get_attribute("id") if rows else None
                        success = True
                        break # Exit retry loop
                    except Exception as pg_err:
                        logger.warning(f"Error parsing page {page_num}: {pg_err}. Retrying ({retry_count+1}/{max_retries})...")
                        self._bump_backoff("page")
                        self._random_sleep(2, 4)
                        self.page.reload()
                        retry_count += 1
                
                if not success:
                    logger.error(f"Failed to parse page {page_num} after retries. Stopping to preserve data.")
                    break

                # Pagination Logic
                if self.browse_every and page_num % self.browse_every == 0:
                    self._simulate_browse_path(stage=f"page-{page_num}")
                if not self._go_next_page(last_first_id):
                    logger.success("Reached last page or pagination blocked.")
                    break

                page_num += 1
                
            # Save Data
            if orders:
                df = pd.DataFrame(orders)
                if "日期" in df.columns:
                    try:
                        df["日期"] = pd.to_datetime(df["日期"], errors="coerce")
                        # 保证同一订单行紧邻，便于后续金额合并
                        sort_cols = ["日期", "订单"] if "订单" in df.columns else ["日期"]
                        sort_order = [False, True] if len(sort_cols) == 2 else [False]
                        df.sort_values(by=sort_cols, ascending=sort_order, inplace=True)
                    except Exception:
                        logger.warning("日期列无法解析为时间，保持原始顺序。")
                split_orders = set(df.loc[df.get("拆单标记") == True, "订单"].tolist()) if "拆单标记" in df.columns else set()
                # 同一订单多商品且未拆单：仅保留首行金额，便于后续合并。
                df = self._collapse_order_amounts(df, split_orders)
                filename = f"jd_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filepath = self.download_dir / filename
                os.makedirs(self.download_dir, exist_ok=True)
                if "拆单标记" in df.columns:
                    df = df.drop(columns=["拆单标记"])
                df.to_excel(filepath, index=False)
                # 合并金额单元格，避免一单多行重复显示。
                try:
                    self._merge_order_amount_cells(filepath, df, split_orders)
                except Exception as merge_err:
                    logger.warning(f"金额单元格合并失败: {merge_err}")
                # Embed images if possible
                if self.embed_images:
                    try:
                        self._embed_images(filepath, df)
                    except Exception as img_err:
                        logger.warning(f"Embed images failed: {img_err}")
                else:
                    logger.info("跳过商品图片嵌入（JD_EMBED_IMAGES=0）。")
                unique_orders = len(set(o["订单"] for o in orders if "订单" in o))
                logger.success(f"Task Completed. Captured {unique_orders} orders ({len(orders)} items). Saved to {filepath}")
                return {
                    "status": "success", 
                    "file": str(filepath), 
                    "count": len(orders), 
                    "order_count": unique_orders
                }
            else:
                return {"status": "empty", "message": "No orders found"}

        except Exception as e:
            logger.error(f"Critical Scraping Error: {e}")
            return {"status": "error", "message": str(e)}
        finally:
            self.close_browser()

    def _parse_row(self, tbody):
        """
        Parse a single order tbody (which may contain multiple products).
        Return list with Chinese字段，一商品一行。
        """
        try:
            tr_th = tbody.query_selector("tr.tr-th")
            if not tr_th:
                return []

            id_el = tr_th.query_selector("a[name='orderIdLinks']")
            if id_el:
                order_id = id_el.inner_text().strip()
            else:
                raw_id = tbody.get_attribute("id") or ""
                order_id = raw_id.replace("tb-order-", "").replace("tb-", "")

            date_el = tr_th.query_selector("span.dealtime")
            order_time = date_el.get_attribute("title") if date_el and date_el.get_attribute("title") else ""
            if not order_time:
                header_text = tr_th.inner_text()
                parts = header_text.strip().split(" ")
                order_time = f"{parts[0]} {parts[1]}" if len(parts) >= 2 else ""

            shop_el = tr_th.query_selector(".shop-name a")
            shop_name = shop_el.inner_text().strip() if shop_el else "自营/未知"

            status_el = tbody.query_selector(".order-status")
            status = status_el.inner_text().strip() if status_el else ""

            # 拆单标记：分拆主体或子单的 tbody 会带 split-tbody 或 data-parentid
            tbody_class = tbody.get_attribute("class") or ""
            is_split = ("split-tbody" in tbody_class) or bool(tbody.get_attribute("data-parentid"))

            # 收货人（列表页直接展示）
            receiver = ""
            consignee_el = tbody.query_selector(".consignee, td.consignee, .consignee a")
            if consignee_el:
                receiver = consignee_el.inner_text().strip()

            # 订单详情链接（优先“订单详情”按钮，其次 fallback 直拼 URL）
            detail_url = ""
            detail_link_el = tbody.query_selector("a:has-text('订单详情')")
            if detail_link_el:
                detail_url = detail_link_el.get_attribute("href") or ""
                if detail_url.startswith("//"):
                    detail_url = "https:" + detail_url
                elif detail_url.startswith("/"):
                    detail_url = urljoin(self.base_url, detail_url)
            if not detail_url and order_id:
                # JD 订单详情页通用格式
                detail_url = f"https://details.jd.com/normal/item.action?orderid={order_id}"

            product_rows = tbody.query_selector_all("tr.tr-bd") or []
            if not product_rows:
                product_rows = [tbody]

            parsed_items = []
            order_address = None
            for row in product_rows:
                # Skip separator rows in split orders
                if "sep-tr-bd" in (row.get_attribute("class") or ""):
                    continue

                name_el = row.query_selector(".p-name a, .p-name em, .p-name")
                product_name = name_el.inner_text().strip() if name_el else ""

                if not product_name:
                    continue

                link = ""
                if name_el:
                    link = name_el.get_attribute("href") or ""
                    if link.startswith("//"):
                        link = "https:" + link

                sku = ""
                sku_el = row.query_selector("[data-sku]") or row.query_selector(".p-sku")
                if sku_el:
                    sku = (sku_el.get_attribute("data-sku") or sku_el.inner_text() or "").strip()
                if not sku and link:
                    import re
                    m = re.search(r"/(\\d+)\\.html", link)
                    if m:
                        sku = m.group(1)

                qty = 1
                num_el = row.query_selector(".goods-number, .goods-number em, .goods-num")
                if num_el:
                    qty_val = self._extract_number(num_el.inner_text())
                    if qty_val:
                        qty = qty_val

                price = ""
                # Fixed selector: target direct span or text, avoiding .ftx-13 (payment method)
                price_el = row.query_selector(".amount span")
                if not price_el:
                     price_el = row.query_selector(".p-price strong")
                if price_el:
                    price = price_el.inner_text().replace("¥", "").strip()

                try:
                    amount_val = float(price) * int(qty)
                except Exception:
                    amount_val = price or ""

                # Fetch order address once per order (detail页)
                if order_address is None:
                    # 先尝试缓存
                    if order_id in self.address_cache:
                        order_address = self.address_cache[order_id]
                    elif detail_url:
                        order_address = self._get_order_address(order_id, detail_url)

                img_src = ""
                img_el = row.query_selector(".p-img img")
                if img_el:
                    # Check lazy load attribute first or fallback to src
                    src = img_el.get_attribute("src") or ""
                    lazy = img_el.get_attribute("data-lazy-img")
                    
                    if lazy and lazy != "done":
                         src = lazy
                    
                    if src.startswith("//"):
                        src = "https:" + src
                    img_src = src

                parsed_items.append({
                    "日期": order_time,
                    "订单": order_id,
                    "商品名称": product_name,
                    "型号": sku,
                    "数量": qty,
                    "下单金额": amount_val,
                    "姓名": receiver,
                    "地址": order_address or "",
                    "店铺": shop_name,
                    "状态": status,
                    "拆单标记": is_split,
                    "商品图片": img_src, # Moved to last
                })

            return parsed_items
        except Exception as e:
            logger.warning(f"Error parsing order {tbody.get_attribute('id')}: {e}")
            return []

    def _extract_number(self, text):
        try:
            import re
            m = re.search(r"\d+", text or "")
            return int(m.group()) if m else None
        except Exception:
            return None

    def _safe_int(self, val, default=0):
        try:
            return int(val)
        except Exception:
            return default

    def _safe_float(self, val, default=0.0):
        try:
            return float(val)
        except Exception:
            return default

    def _parse_browse_urls(self, raw: str):
        if not raw:
            return []
        urls = [u.strip() for u in raw.split(",") if u.strip()]
        return urls

    def _load_or_create_fingerprint(self):
        default_ua = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        data = {}
        if self.fingerprint_file.exists():
            try:
                with open(self.fingerprint_file, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
            except Exception as e:
                logger.warning(f"读取指纹文件失败，将重新生成: {e}")

        def _env_int(name, fallback):
            raw = os.getenv(name)
            if raw:
                try:
                    return int(raw)
                except Exception:
                    pass
            return fallback

        def _env_bool(name, fallback):
            raw = os.getenv(name)
            if raw is None:
                return fallback
            return raw.lower() not in ("0", "false", "no")

        viewport_data = data.get("viewport") or {}
        width = _env_int("JD_VIEWPORT_W", viewport_data.get("width"))
        height = _env_int("JD_VIEWPORT_H", viewport_data.get("height"))
        if not width or not height:
            width = random.randint(1366, 1920)
            height = random.randint(768, 1080)

        fingerprint = {
            "user_agent": os.getenv("JD_UA", data.get("user_agent", default_ua)),
            "locale": os.getenv("JD_LOCALE", data.get("locale", "zh-CN")),
            "timezone_id": os.getenv("JD_TZ", data.get("timezone_id", "Asia/Shanghai")),
            "accept_language": os.getenv("JD_ACCEPT_LANGUAGE", data.get("accept_language", "zh-CN,zh;q=0.9")),
            "viewport": {"width": int(width), "height": int(height)},
            "device_scale_factor": self._safe_float(os.getenv("JD_DEVICE_SCALE", data.get("device_scale_factor", "1.0")), default=1.0),
            "is_mobile": _env_bool("JD_IS_MOBILE", bool(data.get("is_mobile", False))),
        }

        try:
            self.fingerprint_file.parent.mkdir(parents=True, exist_ok=True)
            with open(self.fingerprint_file, "w", encoding="utf-8") as f:
                json.dump(fingerprint, f, ensure_ascii=True, indent=2)
        except Exception as e:
            logger.warning(f"写入指纹文件失败: {e}")

        return fingerprint

    def _collapse_order_amounts(self, df: pd.DataFrame, split_orders: set):
        """同一订单的多商品仅保留首行金额（拆单订单保留各行金额）。"""
        if "订单" not in df.columns or "下单金额" not in df.columns:
            return df
        orders = df["订单"].tolist()
        # 记录第一个出现的位置，后续同订单的金额置空（仅限未拆单）
        first_seen = {}
        for idx, oid in enumerate(orders):
            if oid in split_orders:
                continue
            if oid not in first_seen:
                first_seen[oid] = idx
            else:
                df.at[df.index[idx], "下单金额"] = ""
        return df

    def _merge_order_amount_cells(self, filepath: Path, df: pd.DataFrame, split_orders: set):
        """在 Excel 中按订单合并金额列单元格（仅未拆单订单合并）。"""
        if "订单" not in df.columns or "下单金额" not in df.columns:
            return
        wb = load_workbook(filepath)
        ws = wb.active
        amount_col = df.columns.get_loc("下单金额") + 1  # openpyxl 使用 1-based

        orders = df["订单"].tolist()
        header_offset = 2  # 数据从第 2 行开始（第 1 行是表头）
        # 统计连续区间
        start = 0
        while start < len(orders):
            end = start
            while end + 1 < len(orders) and orders[end + 1] == orders[start]:
                end += 1
            if end > start and orders[start] not in split_orders:
                start_row = header_offset + start
                end_row = header_offset + end
                for r in range(start_row + 1, end_row + 1):
                    ws.cell(row=r, column=amount_col).value = ""
                ws.merge_cells(start_row=start_row, start_column=amount_col, end_row=end_row, end_column=amount_col)
            start = end + 1

        wb.save(filepath)

    def _get_order_address(self, order_id: str, detail_url: str):
        """进入订单详情页提取地址，仅提取一次并缓存。"""
        try:
            if order_id in self.address_cache:
                return self.address_cache[order_id]

            if not detail_url:
                return ""

            target = detail_url
            if target.startswith("//"):
                target = "https:" + target
            elif target.startswith("/"):
                target = urljoin(self.base_url, target)

            detail_page = self._get_detail_page()
            info_text = ""
            detail_ok = False
            try:
                self._rate_limit("detail")
                detail_page.goto(target, wait_until="domcontentloaded", timeout=20000)
                detail_page.wait_for_load_state("networkidle", timeout=12000)
                reason = self._detect_risk_page(detail_page)
                if reason and self._handle_risk_page(detail_page, reason, fatal=False):
                    self._reset_detail_page()
                    return ""
                if random.random() < self.detail_browse_prob:
                    self._dwell_and_scroll(detail_page, min_s=1.0, max_s=2.8)
                # 等待地址区域渲染（容忍动态加载）
                try:
                    detail_page.wait_for_selector(".item .label, .addr, .info-rcol", timeout=8000)
                except TimeoutError:
                    logger.warning(f"订单详情未及时加载地址元素: {order_id}")

                # 优先：label 含“地址”或“收货地址”，寻找同级 info-rcol
                label_locator = detail_page.locator("span.label").filter(has_text=re.compile("地址"))
                if label_locator.count() > 0:
                    info_text = label_locator.first.evaluate(
                        """el => {
                            const container = el.closest('.item') || el.parentElement;
                            const info = container ? container.querySelector('.info-rcol') : null;
                            return info ? info.textContent.trim() : '';
                        }"""
                    ) or ""

                # 备用：页面上出现的 info-rcol 或 .addr 文本
                if not info_text:
                    fallback = detail_page.locator(".info-rcol, .addr").first
                    if fallback.count() > 0:
                        info_text = (fallback.inner_text() or "").strip()

                # 再次备用：全局查找包含“地址”字段的文案，取冒号后文字
                if not info_text:
                    info_text = detail_page.evaluate(
                        """() => {
                            const textNodes = Array.from(document.querySelectorAll('body *'))
                                .map(el => el.textContent ? el.textContent.trim() : '')
                                .filter(t => t && /地址/.test(t));
                            if (!textNodes.length) return '';
                            const cand = textNodes.find(t => t.length < 200) || textNodes[0];
                            const parts = cand.split(/[:：]/);
                            return parts.length > 1 ? parts.slice(1).join(':').trim() : cand;
                        }"""
                    ) or ""

                # 清洗多余换行与空白
                info_text = re.sub(r"\s+", " ", info_text).strip()

                self.address_cache[order_id] = info_text
                detail_ok = True
                return info_text
            finally:
                try:
                    if not info_text:
                        logger.warning(f"订单地址为空，可能页面结构变化或需要登录态验证：{order_id}")
                finally:
                    if detail_ok:
                        self._decay_backoff("detail")
        except Exception as e:
            logger.warning(f"获取订单地址失败 {order_id}: {e}")
            self._bump_backoff("detail")
            self._reset_detail_page()
            self.address_cache[order_id] = ""
            return ""

    def _embed_images(self, filepath, df):
        """
        Download images from '商品图片' column and embed into Excel file.
        Uses a temp file to avoid corrupting the main file on failure.
        """
        if "商品图片" not in df.columns:
            return

        tmp_path = Path(filepath).with_suffix(".tmp.xlsx")
        wb = load_workbook(filepath)
        ws = wb.active

        col_idx = list(df.columns).index("商品图片") + 1  # 1-based
        col_letter = get_column_letter(col_idx)

        headers = {
            "User-Agent": random.choice(self.user_agents),
            "Accept-Language": self.accept_language,
            "Referer": "https://www.jd.com/",
        }

        for idx, url in enumerate(df["商品图片"]):
            if not url:
                continue
            excel_row = idx + 2  # header is row 1
            try:
                img_bytes = self._fetch_image_bytes(url, headers)
                img_bytes = io.BytesIO(img_bytes)
                img = XLImage(img_bytes)
                img.width = 80
                img.height = 80
                cell_addr = f"{col_letter}{excel_row}"
                ws.add_image(img, cell_addr)
                ws[cell_addr].value = ""
            except Exception as e:
                logger.warning(f"Embed image failed for row {excel_row}: {e}")
                continue

        wb.save(tmp_path)
        Path(tmp_path).replace(filepath)

    def _fetch_image_bytes(self, url: str, headers: dict):
        last_err = None
        for attempt in range(1, self.image_retries + 2):
            try:
                if url.startswith("//"):
                    url = "https:" + url
                self._rate_limit("image")
                if self.context and not self.context.is_closed():
                    resp = self.context.request.get(url, headers=headers, timeout=15000)
                    try:
                        status = resp.status
                        if status in (403, 429):
                            last_err = Exception(f"image status {status}")
                            self._bump_backoff("image", factor=1.8)
                            self._random_sleep(0.8, 1.6)
                            continue
                        if status >= 400:
                            raise Exception(f"image status {status}")
                        body = resp.body()
                    finally:
                        try:
                            resp.dispose()
                        except Exception:
                            pass
                    self._decay_backoff("image")
                    return body

                if not self.http:
                    self.http = requests.Session()
                resp = self.http.get(url, headers=headers, timeout=15)
                if resp.status_code in (403, 429):
                    last_err = Exception(f"image status {resp.status_code}")
                    self._bump_backoff("image", factor=1.8)
                    self._random_sleep(0.8, 1.6)
                    continue
                resp.raise_for_status()
                self._decay_backoff("image")
                return resp.content
            except Exception as e:
                last_err = e
                self._bump_backoff("image", factor=1.4)
                self._random_sleep(0.6, 1.2)
        if last_err:
            raise last_err
        raise Exception("image download failed")

    def _wait_for_orders_ready(self):
        """Ensure the order table and rows are present before parsing."""
        self.page.wait_for_selector("table.order-tb", timeout=8000)
        self.page.wait_for_selector("tbody[id^='tb-']", timeout=8000)
        if "passport.jd.com" in self.page.url:
            raise Exception("会话失效，请重新登录。")
        reason = self._detect_risk_page(self.page)
        if reason:
            self._handle_risk_page(self.page, reason, fatal=True)

    def _goto_with_retry(self, url: str, wait_until="domcontentloaded", retries: int = None, timeout: int = 20000):
        """Navigate with basic retry to handle临时 DNS/网络抖动."""
        retries = retries or self.goto_retries
        last_err = None
        for attempt in range(1, retries + 1):
            try:
                self._rate_limit("page")
                self.page.goto(url, wait_until=wait_until, timeout=timeout)
                reason = self._detect_risk_page(self.page)
                if reason:
                    self._handle_risk_page(self.page, reason, fatal=True)
                self._decay_backoff("page")
                logger.info(f"Goto success [{attempt}/{retries}]: {url}")
                return True
            except Exception as e:
                last_err = e
                logger.warning(f"Goto失败({attempt}/{retries}): {url} -> {e}")
                self._bump_backoff("page")
                self._random_sleep(1.2, 2.5)
        if last_err:
            raise last_err

    def _go_next_page(self, last_first_id: str):
        """Handle pagination robustly; return False when no more pages."""
        candidate_selectors = [
            "a.ui-pager-next",        # old selector
            "div.pagin a.next",       # JD order list selector
            "a.next"                  # fallback
        ]

        next_locator = None
        for sel in candidate_selectors:
            loc = self.page.locator(sel).first
            if loc.count() > 0:
                next_locator = loc
                break

        if not next_locator:
            logger.info("No next-page control found; assuming last page.")
            return False

        classes = (next_locator.get_attribute("class") or "").lower()
        if "disabled" in classes or "ui-pager-disabled" in classes:
            return False

        href = (next_locator.get_attribute("href") or "").strip()
        self._random_sleep(1.2, 3.5)

        try:
            if href and href != "#" and "javascript" not in href.lower():
                # Normalize protocol-relative URLs
                target = f"https:{href}" if href.startswith("//") else urljoin(self.page.url, href)
                self.page.goto(target, wait_until="domcontentloaded")
            else:
                with self.page.expect_navigation(wait_until="domcontentloaded", timeout=12000):
                    next_locator.click()
        except TimeoutError as e:
            logger.warning(f"Pagination navigation timeout: {e}")
            return False

        # Wait for content change; JD may be ajax or full navigation.
        try:
            self.page.wait_for_load_state("networkidle", timeout=12000)
        except TimeoutError:
            logger.warning("Network idle wait timed out, checking DOM change directly.")

        try:
            self.page.wait_for_selector("tbody[id^='tb-']", timeout=8000)
            self.page.wait_for_function(
                """(firstId) => {
                    const first = document.querySelector("tbody[id^='tb-']");
                    return !firstId || (first && first.id !== firstId);
                }""",
                arg=last_first_id,
                timeout=12000
            )
        except TimeoutError:
            logger.warning("Pagination DOM did not change after navigating.")

        new_first_el = self.page.query_selector("tbody[id^='tb-']")
        new_first_id = new_first_el.get_attribute("id") if new_first_el else None
        
        if last_first_id and new_first_id == last_first_id:
            logger.warning("Still on the same page after attempting next; stopping to avoid infinite loop.")
            return False

        return True
